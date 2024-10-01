from io import BytesIO

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.worksheet import Worksheet


def write_worksheet(writer: pd.ExcelWriter, name: str, df: pd.DataFrame, format: bool = True):
    """Write a pandas dataframe to an excel sheet.

    Args:
        writer (pd.ExcelWriter): the writer instance
        name (str): the worksheet name
        df (pd.DataFrame): the dataframe to write
        format (bool, default True): style the worksheet
    """
    # Remove timezone from datetime objects to make Excel compatible
    df = df.copy()
    for col in df.select_dtypes(include="datetimetz").columns:
        df[col] = df[col].dt.tz_localize(None)

    try:
        df.to_excel(writer, sheet_name=name, index=False, freeze_panes=(1, 0))
    except IllegalCharacterError:
        # Clean data frame to remove illegal characters, such as "\x02", inplace.
        columns: list[str] = df.select_dtypes(include="object").columns.values
        for column in columns:
            if hasattr(df.loc[:, column], "str"):
                df.loc[:, column] = df.loc[:, column].str.replace(
                    ILLEGAL_CHARACTERS_RE, "", regex=True
                )
        df.to_excel(writer, sheet_name=name, index=False, freeze_panes=(1, 0))

    if format:
        format_worksheet(writer.sheets[name])


def format_worksheet(ws: Worksheet):
    """Format the worksheet with additional layout and styling.

    Args:
        ws (Worksheet): a Worksheet
    """
    # enable filters
    ws.auto_filter.ref = ws.dimensions
    # fill header
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="1F497D")
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal="left")
    # resize columns
    for idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 10


def get_writer() -> (BytesIO, pd.ExcelWriter):
    """Standard constructor to generate a bytes stream and an excel writer"""
    f = BytesIO()
    writer = pd.ExcelWriter(f, date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD HH:MM:SS")
    return (f, writer)
