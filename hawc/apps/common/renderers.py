import json
from io import BytesIO, StringIO

import matplotlib.pyplot as plt
import pandas as pd
from django.conf import settings
from django.utils.text import slugify
from matplotlib.axes import Axes
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from rest_framework import status
from rest_framework.renderers import BaseRenderer, BrowsableAPIRenderer
from rest_framework.response import Response

from .helper import FlatExport, ReportExport, rename_duplicate_columns


class DocxRenderer(BaseRenderer):
    """
    Renders a ReportExport object into a docx file.
    """

    media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    format = ""

    def render(self, data, accepted_media_type=None, renderer_context=None):
        # return error or OPTIONS as JSON
        status_code = renderer_context["response"].status_code
        method = renderer_context["request"].method if "request" in renderer_context else None

        # throw error if we don't have a ReportExport
        if not isinstance(data, ReportExport):
            success = status.is_success(status_code)
            if (method == "OPTIONS" or not success) and isinstance(data, dict):
                return json.dumps(data)
            raise ValueError(f"Expecting `ReportExport`; got {type(data)}")

        file = BytesIO()
        data.docx.save(file)
        response = renderer_context["response"]
        response["Content-Disposition"] = f"attachment; filename={data.filename}.docx"
        return file.getvalue()


class SvgRenderer(BaseRenderer):
    media_type = "image/svg+xml"
    format = "svg"

    def render(self, ax: Axes, accepted_media_type=None, renderer_context=None):
        if isinstance(ax, dict):
            return f'<svg xmlns="http://www.w3.org/2000/svg"><text y="15">{json.dumps(ax)}</text></svg>'
        f = StringIO()
        ax.figure.savefig(f, format="svg")
        plt.close(ax.figure)
        return f.getvalue()


class PandasBaseRenderer(BaseRenderer):
    """
    Renders DataFrames using their built in pandas implementation.
    Borrowed heavily from `django-rest-pandas`.
    """

    def render(self, data, accepted_media_type=None, renderer_context=None):
        # return error or OPTIONS as JSON
        status_code = renderer_context["response"].status_code
        method = renderer_context["request"].method if "request" in renderer_context else None
        if not status.is_success(status_code) or method == "OPTIONS":
            if isinstance(data, dict):
                return json.dumps(data)
            else:
                raise ValueError(f"Expecting data as `dict`; got {type(data)}")

        # throw error if we don't have a FlatExport
        if not isinstance(data, FlatExport):
            raise ValueError(f"Expecting `FlatExport`; got {type(data)}")

        return self.render_dataframe(data, renderer_context["response"])


class PandasHtmlRenderer(PandasBaseRenderer):
    """
    Renders dataframe as Html
    """

    media_type = "text/html"
    format = "html"

    def render_dataframe(self, export: FlatExport, response: Response) -> str:
        with pd.option_context("display.max_colwidth", None):
            return export.df.fillna("-").to_html(index=False)


class PandasCsvRenderer(PandasBaseRenderer):
    """
    Renders dataframe as CSV
    """

    media_type = "text/csv"
    format = "csv"

    def render_dataframe(self, export: FlatExport, response: Response) -> str:
        # set line terminator to keep consistent on windows too
        return export.df.to_csv(index=False, lineterminator="\n")


class PandasTsvRenderer(PandasBaseRenderer):
    """
    Renders dataframe as TSV
    """

    media_type = "text/tab-separated-values"
    format = "tsv"

    def render_dataframe(self, export: FlatExport, response: Response) -> str:
        # set line terminator to keep consistent on windows too
        return export.df.to_csv(index=False, sep="\t", lineterminator="\n")


class PandasJsonRenderer(PandasBaseRenderer):
    """
    Renders dataframe as JSON
    """

    media_type = "application/json"
    format = "json"

    def render_dataframe(self, export: FlatExport, response: Response) -> str:
        if export.df.columns.has_duplicates:
            rename_duplicate_columns(export.df)
        return export.df.to_json(orient="records")


class PandasBrowsableAPIRenderer(BrowsableAPIRenderer):
    """
    Renders dataframe using the DRF browser.

    This can be useful in debugging to view the django debug toolbar
    for database query performance.
    """

    def get_content(self, renderer, data, accepted_media_type, renderer_context):
        # handle OPTIONS
        if isinstance(data, dict):
            return json.dumps(data, indent=4)

        # handle dataframe
        if data.df.columns.has_duplicates:
            rename_duplicate_columns(data.df)
        return data.df.to_json(orient="records", indent=4)


class PandasXlsxBinaryRenderer(PandasBaseRenderer):
    """
    Renders dataframe as xlsx
    """

    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "xlsx"

    def render(self, data: BytesIO, accepted_media_type=None, renderer_context=None):
        # return error or OPTIONS as JSON
        status_code = renderer_context["response"].status_code
        method = renderer_context["request"].method if "request" in renderer_context else None
        if not status.is_success(status_code) or method == "OPTIONS":
            if isinstance(data, dict):
                return json.dumps(data)
            else:
                raise ValueError(f"Expecting data as `dict`; got {type(data)}")
        return data.getvalue()


class PandasXlsxRenderer(PandasBaseRenderer):
    """
    Renders dataframe as xlsx
    """

    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "xlsx"

    def render_dataframe(self, export: FlatExport, response: Response) -> bytes:
        response["Content-Disposition"] = f"attachment; filename={slugify(export.filename)}.xlsx"

        f = BytesIO()
        with pd.ExcelWriter(
            f, date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD HH:MM:SS"
        ) as writer:
            write_worksheet(writer, export.df, "data")
            if isinstance(export.metadata, pd.DataFrame):
                write_worksheet(writer, export.metadata, "metadata")
            format_xlsx(writer)

        return f.getvalue()


def write_worksheet(writer: pd.ExcelWriter, df: pd.DataFrame, name: str):
    # Remove timezone from datetime objects to make Excel compatible
    df = df.copy()
    for col in df.select_dtypes(include="datetimetz").columns:
        df[col] = df[col].dt.tz_localize(None)

    try:
        df.to_excel(writer, sheet_name=name, index=False, freeze_panes=(1, 0))
    except IllegalCharacterError:
        # Clean data frame to remove illegal characters, such as "\x02", inplace.
        columns: list[str] = df.select_dtypes(include="object").columns.values
        for column in columns:
            if hasattr(df.loc[:, column], "str"):
                df.loc[:, column] = df.loc[:, column].str.replace(
                    ILLEGAL_CHARACTERS_RE, "", regex=True
                )
        df.to_excel(writer, sheet_name=name, index=False, freeze_panes=(1, 0))


def format_xlsx(writer: pd.ExcelWriter):
    # enable filters
    for ws in writer.book.worksheets:
        # enable filters
        ws.auto_filter.ref = ws.dimensions
        # fill header
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="1F497D")
                cell.font = Font(color="FFFFFF")
                cell.alignment = Alignment(horizontal="left")
        # resize columns
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 10


PandasRenderers = [
    PandasJsonRenderer,
    PandasHtmlRenderer,
    PandasCsvRenderer,
    PandasTsvRenderer,
    PandasXlsxRenderer,
]

if settings.DEBUG:
    # insert at position 1 to keep JSON the default renderer
    PandasRenderers.insert(1, PandasBrowsableAPIRenderer)
