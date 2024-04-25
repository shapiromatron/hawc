import json
from pathlib import Path
from urllib.parse import urlparse, urlunparse
from zipfile import BadZipFile

import pandas as pd
import plotly.io as pio
from django import forms
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..animal.autocomplete import EndpointAutocomplete
from ..animal.forms import MultipleEndpointChoiceField
from ..animal.models import Endpoint
from ..assessment.models import DoseUnits
from ..common import validators
from ..common.autocomplete import AutocompleteChoiceField
from ..common.clean import sanitize_html
from ..common.dynamic_forms import Schema
from ..common.forms import (
    BaseFormHelper,
    CopyForm,
    DynamicFormField,
    QuillField,
    check_unique_for_assessment,
)
from ..common.helper import new_window_a
from ..common.validators import validate_html_tags, validate_hyperlinks, validate_json_pydantic
from ..lit.models import ReferenceFilterTag
from ..study.autocomplete import StudyAutocomplete
from . import autocomplete, constants, models, prefilters


class SummaryTextForm(forms.ModelForm):
    class Meta:
        model = models.SummaryText
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        kwargs.pop("parent", None)
        super().__init__(*args, **kwargs)


class SummaryTableForm(forms.ModelForm):
    class Meta:
        model = models.SummaryTable
        exclude = ("assessment", "table_type")
        field_classes = {"caption": QuillField}

    def __init__(self, *args, **kwargs):
        self.assessment = kwargs.pop("parent", None)
        table_type = kwargs.pop("table_type", None)
        super().__init__(*args, **kwargs)
        if not self.instance.id:
            self.instance = models.SummaryTable.build_default(self.assessment.id, table_type)
        # TODO - we shouldn't have to do this; and 400 errors are not rendering
        # instead, switch to htmx or use a REST API
        if self.initial:
            self.instance.title = self.initial["title"]
            self.instance.slug = self.initial["slug"]
            self.instance.content = self.initial["content"]
            self.instance.published = self.initial["published"]
            self.instance.caption = self.initial["caption"]

    def clean_slug(self):
        return check_unique_for_assessment(self, "slug")

    def clean_title(self):
        return check_unique_for_assessment(self, "title")


class SummaryTableSelectorForm(forms.Form):
    table_type = forms.TypedChoiceField(coerce=int, choices=constants.TableType.choices)

    def __init__(self, *args, **kwargs):
        self.assessment = kwargs.pop("parent")
        _ = kwargs.pop("instance")
        super().__init__(*args, **kwargs)

    @property
    def helper(self):
        url = models.SummaryTable.get_list_url(self.assessment.id)
        return BaseFormHelper(
            self,
            legend_text="Select table type",
            help_text="""
            HAWC has a number of predefined table formats which are designed for different applications
            of the tool. Please select the table type you wish to create.
            """,
            submit_text="Create table",
            cancel_url=url,
        )


class SummaryTableModelChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.assessment}: [{obj.get_table_type_display()}] {obj}"


class SummaryTableCopySelectorForm(CopyForm):
    legend_text = "Copy summary table"
    help_text = "Select an existing summary table as a template to create a new one."
    create_url_pattern = "summary:tables_create"
    selector = forms.ModelChoiceField(
        queryset=models.SummaryTable.objects.all(), empty_label=None, label="Select template"
    )

    def __init__(self, *args, **kw):
        self.user = kw.pop("user")
        super().__init__(*args, **kw)
        self.fields["selector"].queryset = (
            models.SummaryTable.objects.clonable_queryset(self.user)
            .select_related("assessment")
            .order_by("assessment", "title")
        )
        self.fields["selector"].label_from_instance = (
            lambda obj: f"{obj.assessment} | {{{obj.get_table_type_display()}}} | {obj}"
        )

    def get_success_url(self):
        table = self.cleaned_data["selector"]
        return (
            reverse(self.create_url_pattern, args=(self.parent.id, table.table_type))
            + f"?initial={table.id}"
        )

    def get_cancel_url(self):
        return reverse("summary:tables_list", args=(self.parent.id,))


class VisualForm(forms.ModelForm):
    class Meta:
        model = models.Visual
        exclude = ("assessment", "visual_type", "evidence_type", "prefilters")

    def __init__(self, *args, **kwargs):
        assessment = kwargs.pop("parent", None)
        visual_type = kwargs.pop("visual_type", None)
        evidence_type = kwargs.pop("evidence_type", None)
        super().__init__(*args, **kwargs)
        if "settings" in self.fields:
            self.fields["settings"].widget.attrs["rows"] = 2
        if assessment:
            self.instance.assessment = assessment
        if visual_type is not None:  # required if value is 0
            self.instance.visual_type = visual_type
        if self.instance.visual_type not in [
            constants.VisualType.BIOASSAY_AGGREGATION,
            constants.VisualType.ROB_HEATMAP,
            constants.VisualType.LITERATURE_TAGTREE,
            constants.VisualType.EXTERNAL_SITE,
            constants.VisualType.EXPLORE_HEATMAP,
            constants.VisualType.PLOTLY,
            constants.VisualType.IMAGE,
        ]:
            self.fields["sort_order"].widget = forms.HiddenInput()
        if self.instance.id is None:
            self.instance.evidence_type = evidence_type

    def setHelper(self):
        for fld in list(self.fields.keys()):
            widget = self.fields[fld].widget
            if type(widget) != forms.CheckboxInput:
                widget.attrs["class"] = "col-md-12"

        if self.instance.id:
            inputs = {
                "legend_text": f"Update {self.instance}",
                "help_text": "Update an existing visualization.",
                "cancel_url": self.instance.get_absolute_url(),
            }
        else:
            inputs = {
                "legend_text": "Create new visualization",
                "help_text": """
                    Create a custom-visualization.
                    Generally, you will select a subset of available data on the
                    "Data" tab, then will customize the visualization using the
                    "Settings" tab. To view a preview of the visual at any time,
                    select the "Preview" tab.
                """,
                "cancel_url": self.instance.get_list_url(self.instance.assessment_id),
            }

        helper = BaseFormHelper(self, **inputs)

        helper.form_id = "visualForm"
        return helper

    def clean_slug(self):
        return check_unique_for_assessment(self, "slug")

    def clean_title(self):
        return check_unique_for_assessment(self, "title")

    def clean_caption(self):
        caption = self.cleaned_data["caption"]
        validators.validate_hyperlinks(caption)
        return sanitize_html.clean_html(caption)

    def clean_evidence_type(self):
        visual_type = self.cleaned_data["visual_type"]
        evidence_type = self.cleaned_data["evidence_type"]
        if evidence_type not in constants.VISUAL_EVIDENCE_CHOICES[visual_type]:
            raise forms.ValidationError(
                f"Invalid evidence type {evidence_type} for visual {visual_type}."
            )


class VisualModelChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.assessment}: [{obj.get_visual_type_display()}] {obj}"


class VisualSelectorForm(CopyForm):
    legend_text = "Copy visualization"
    help_text = "Select an existing visualization from this assessment to copy as a template for a new one. This will include all model-settings, and the selected dataset."
    create_url_pattern = "summary:visualization_create"
    selector = forms.ModelChoiceField(
        queryset=models.Visual.objects.all(), empty_label=None, label="Select template"
    )

    def __init__(self, *args, **kw):
        queryset = kw.pop("queryset")
        super().__init__(*args, **kw)
        self.fields["selector"].queryset = queryset.order_by("assessment", "title")
        self.fields["selector"].label_from_instance = lambda obj: f"{obj.assessment} | {obj}"

    def get_success_url(self):
        visual = self.cleaned_data["selector"]
        return (
            reverse("summary:visualization_create", args=(self.parent.id, visual.visual_type))
            + f"?initial={visual.pk}"
        )

    def get_cancel_url(self):
        return reverse("summary:visualization_list", args=(self.parent.id,))


class EndpointAggregationForm(VisualForm):
    endpoints = MultipleEndpointChoiceField(required=True, queryset=Endpoint.objects.none())

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.helper = self.setHelper()
        self.helper.attrs["novalidate"] = ""
        self.fields["dose_units"].queryset = DoseUnits.objects.get_animal_units(
            self.instance.assessment
        )
        self.fields["endpoints"].widget.attrs["size"] = 20
        self.fields["endpoints"].queryset = Endpoint.objects.assessment_qs(
            self.instance.assessment
        ).selector()

    class Meta:
        model = models.Visual
        exclude = (
            "assessment",
            "visual_type",
            "evidence_type",
            "settings",
            "prefilters",
            "studies",
            "sort_order",
            "image",
        )


class CrossviewForm(VisualForm):
    def _get_prefilter_form(self, data, **form_kwargs):
        prefix = form_kwargs.pop("prefix", None)
        prefilter = self.prefilter_cls(
            data=data, prefix=prefix, assessment=self.instance.assessment, form_kwargs=form_kwargs
        )
        form = prefilter.form
        prefilter.set_form_options(form)
        return form

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["dose_units"].queryset = DoseUnits.objects.get_animal_units(
            self.instance.assessment
        )
        self.prefilter_cls = prefilters.get_prefilter_cls(
            self.instance.visual_type, self.instance.evidence_type, self.instance.assessment
        )
        self.fields["prefilters"] = DynamicFormField(
            prefix="prefilters", form_class=self._get_prefilter_form, label=""
        )
        self.helper = self.setHelper()

    class Meta:
        model = models.Visual
        exclude = ("assessment", "visual_type", "evidence_type", "endpoints", "studies", "image")


class RoBForm(VisualForm):
    def _get_prefilter_form(self, data, **form_kwargs):
        prefix = form_kwargs.pop("prefix", None)
        prefilter = self.prefilter_cls(
            data=data, prefix=prefix, assessment=self.instance.assessment, form_kwargs=form_kwargs
        )
        form = prefilter.form
        prefilter.set_form_options(form)
        return form

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.prefilter_cls = prefilters.get_prefilter_cls(
            self.instance.visual_type, self.instance.evidence_type, self.instance.assessment
        )
        self.fields["prefilters"] = DynamicFormField(
            prefix="prefilters", form_class=self._get_prefilter_form, label=""
        )
        self.helper = self.setHelper()

    class Meta:
        model = models.Visual
        exclude = ("assessment", "visual_type", "evidence_type", "dose_units", "endpoints", "image")


class TagtreeForm(VisualForm):
    root_node = forms.TypedChoiceField(
        coerce=int,
        choices=[],
        label="Root node",
        help_text="Select the root node for the tag tree (the left-most node to be displayed)",
        required=True,
    )
    required_tags = forms.TypedMultipleChoiceField(
        coerce=int,
        choices=[],
        label="Filter references by tags",
        help_text="Filter which references are displayed by selecting required tags. If tags are selected, only references which have one or more of these tags (not including descendants) will be displayed.<br><br><i>This field is optional; if no tags are selected, all references will be displayed.</i>",
        required=False,
    )
    pruned_tags = forms.TypedMultipleChoiceField(
        coerce=int,
        choices=[],
        label="Hidden tags",
        help_text="Select tags which should be hidden from the tagtree. If a parent-tag is selected, all child-tags will also be hidden.<br><br><i>This field is optional; if no tags are selected, then all tags will be displayed.</i>",
        required=False,
    )
    hide_empty_tag_nodes = forms.BooleanField(
        label="Hide tags with no references",
        help_text="Prune tree; show only tags which contain at least one reference.",
        required=False,
    )
    width = forms.IntegerField(
        label="Width",
        initial=1280,
        min_value=500,
        max_value=2000,
        required=True,
        help_text="The width of the visual, in pixels.",
    )
    height = forms.IntegerField(
        label="Height",
        initial=800,
        min_value=500,
        max_value=3000,
        required=True,
        help_text="The height of the visual, in pixels. If you have overlapping nodes, add more height",
    )
    show_legend = forms.BooleanField(
        label="Show Legend",
        help_text="Describes what each node type indicates",
        initial=True,
        required=False,
    )
    show_counts = forms.BooleanField(
        label="Show Node Counts",
        help_text="Display the count for each node and scale size accordingly",
        initial=True,
        required=False,
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.helper = self.setHelper()
        self.helper.add_row("root_node", 3, "col-md-4")
        self.helper.add_row(
            "hide_empty_tag_nodes", 5, ["col-md-3", "col-md-2", "col-md-2", "col-md-3", "col-md-2"]
        )

        choices = [
            (tag.id, tag.get_nested_name())
            for tag in ReferenceFilterTag.get_assessment_qs(
                self.instance.assessment_id, include_root=True
            )
        ]
        self.fields["root_node"].choices = choices
        self.fields["required_tags"].choices = choices[1:]
        self.fields["pruned_tags"].choices = choices[1:]

        self.fields["required_tags"].widget.attrs.update(size=10)
        self.fields["pruned_tags"].widget.attrs.update(size=10)

        data = self.instance.settings
        if "root_node" in data:
            self.fields["root_node"].initial = data["root_node"]
        if "required_tags" in data:
            self.fields["required_tags"].initial = data["required_tags"]
        if "pruned_tags" in data:
            self.fields["pruned_tags"].initial = data["pruned_tags"]
        if "hide_empty_tag_nodes" in data:
            self.fields["hide_empty_tag_nodes"].initial = data["hide_empty_tag_nodes"]
        if "width" in data:
            self.fields["width"].initial = data["width"]
        if "height" in data:
            self.fields["height"].initial = data["height"]
        if "show_legend" in data:
            self.fields["show_legend"].initial = data["show_legend"]
        if "show_counts" in data:
            self.fields["show_counts"].initial = data["show_counts"]

    def save(self, commit=True):
        self.instance.settings = dict(
            root_node=self.cleaned_data["root_node"],
            required_tags=self.cleaned_data["required_tags"],
            pruned_tags=self.cleaned_data["pruned_tags"],
            hide_empty_tag_nodes=self.cleaned_data["hide_empty_tag_nodes"],
            width=self.cleaned_data["width"],
            height=self.cleaned_data["height"],
            show_legend=self.cleaned_data["show_legend"],
            show_counts=self.cleaned_data["show_counts"],
        )
        return super().save(commit)

    class Meta:
        model = models.Visual
        fields = (
            "title",
            "slug",
            "caption",
            "published",
            "root_node",
            "required_tags",
            "pruned_tags",
            "hide_empty_tag_nodes",
            "width",
            "height",
            "show_legend",
            "show_counts",
        )


class ExternalSiteForm(VisualForm):
    external_url = forms.URLField(
        label="External URL",
        help_text="""
        <p class="form-text text-muted">
            Embed an external website. The following websites can be linked to:
        </p>
        <ul class="form-text text-muted">
            <li><a href="https://public.tableau.com/">Tableau (public)</a> - press the "share" icon and then select the URL in the "link" text box. For example, <code>https://public.tableau.com/views/PFAS150EpidemiologySQ/EpiOverview</code>.</li>
        </ul>
        <p class="form-text text-muted">
            If you'd like to link to another website, please contact us.
        </p>
        """,
    )
    filters = forms.CharField(
        label="Data filters",
        initial="[]",
        validators=[validate_json_pydantic(constants.TableauFilterList)],
        help_text="""<p class="form-text text-muted">
        Data are expected to be in JSON format, where the each key is a filter name and value is a filter value.
        For more details, view the <a href="https://help.tableau.com/current/api/embedding_api/en-us/docs/embedding_api_filter.html">Tableau documentation</a>. For example: <code>[{"field": "Category", "value": "Technology"}, {"field": "State", "value": "North Carolina,Virginia"}]</code>. To remove filters, set string to <code>[]</code>.
        </p>""",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        data = self.instance.settings
        if "external_url" in data:
            self.fields["external_url"].initial = data["external_url"]
        if "filters" in data:
            self.fields["filters"].initial = json.dumps(data["filters"])

        self.helper = self.setHelper()

    def save(self, commit=True):
        self.instance.settings = dict(
            external_url=self.cleaned_data["external_url"],
            external_url_hostname=self.cleaned_data["external_url_hostname"],
            external_url_path=self.cleaned_data["external_url_path"],
            external_url_query_args=self.cleaned_data["external_url_query_args"],
            filters=json.loads(self.cleaned_data["filters"]),
        )
        return super().save(commit)

    class Meta:
        model = models.Visual
        fields = (
            "title",
            "slug",
            "caption",
            "published",
        )

    DOMAIN_TABLEAU = "public.tableau.com"
    VALID_DOMAINS = {
        DOMAIN_TABLEAU,
    }

    def clean_external_url(self):
        external_url = self.cleaned_data.get("external_url")
        url = urlparse(external_url)

        # check allowlist
        if url.netloc not in self.VALID_DOMAINS:
            msg = f"{url.netloc} not on the list of accepted domains, please contact webmasters to request additions"
            raise forms.ValidationError(msg)

        external_url = urlunparse(("https", url.netloc, url.path, "", "", ""))
        external_url_hostname = urlunparse(("https", url.netloc, "", "", "", ""))

        if url.path == "" or url.path == "/":
            raise forms.ValidationError("A URL path must be specified.")

        external_url_query_args = []
        if url.netloc == self.DOMAIN_TABLEAU:
            external_url_query_args = [":showVizHome=no", ":embed=y"]

        self.cleaned_data.update(
            external_url=external_url,
            external_url_hostname=external_url_hostname,
            external_url_path=url.path,
            external_url_query_args=external_url_query_args,
        )

        return external_url


class ExploreHeatmapForm(VisualForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.helper = self.setHelper()

    class Meta:
        model = models.Visual
        fields = (
            "title",
            "slug",
            "settings",
            "caption",
            "published",
        )


class PlotlyVisualForm(VisualForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["settings"].label = "JSON configuration"
        self.fields[
            "settings"
        ].help_text = f"""Create a {new_window_a("https://plotly.com/", "Plotly")} visual using Python or R, and then export the visual and display to JSON ({new_window_a("https://github.com/plotly/plotly.R/issues/590#issuecomment-220864613" ,"R")} or {new_window_a("https://plotly.github.io/plotly.py-docs/generated/plotly.io.to_json.html", "Python")})."""
        self.helper = self.setHelper()

    class Meta:
        model = models.Visual
        fields = (
            "title",
            "slug",
            "settings",
            "caption",
            "published",
        )

    def clean_settings(self):
        # we remove <extra> tag; by default it's included in plotly visuals but it doesn't pass
        # our html validation checks
        settings: str = (
            json.dumps(self.cleaned_data.get("settings", ""))
            .strip()
            .replace("<extra>", "")
            .replace("</extra>", "")
        )
        try:
            json.loads(settings)
        except ValueError as err:
            raise forms.ValidationError("Invalid JSON format") from err
        if len(json.loads(settings)) == 0:
            raise forms.ValidationError("Settings cannot be empty")
        validate_html_tags(settings)
        validate_hyperlinks(settings)
        try:
            pio.from_json(settings)
        except ValueError as err:
            raise forms.ValidationError("Invalid Plotly figure") from err
        return json.loads(settings)


class PrismaVisualForm(VisualForm):
    class Meta:
        model = models.Visual
        fields = (
            "title",
            "slug",
            "settings",
            "caption",
            "published",
        )


class ImageVisualForm(VisualForm):
    settings_schema = {
        "fields": [
            {
                "name": "alt_text",
                "type": "char",
                "label": "Image alt-text",
                "help_text": "Alternative text if an image fails to display and for accessibility support.",
                "widget": "textarea",
                "css_class": "col-4",
            },
            {
                "name": "max_width",
                "type": "integer",
                "label": "Image maximum width",
                "help_text": "Max width of the image (in pixels). The image will always shrink to be visible in your browser, but if unset, image will grow to be the full size of your browser, which can be huge for high resolution uploads.",
                "initial": 1000,
                "css_class": "col-4",
                "required": False,
            },
        ]
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["image"].required = True
        self.fields["settings"] = DynamicFormField(
            "settings", Schema.model_validate(self.settings_schema).to_form, label=""
        )
        self.helper = self.setHelper()

    def clean_image(self):
        image = self.cleaned_data["image"]
        suffix = Path(image.name).suffix.lower()
        suffixes = (".jpeg", ".jpg", ".png")
        if suffix not in suffixes:
            raise forms.ValidationError(f"File extension must be one of {', '.join(suffixes)}.")
        size_mb = image.size / 1024 / 1024
        if size_mb < 0.01 or size_mb > 3:
            raise forms.ValidationError("Image must be >10KB and <3 MB in size.")
        return image

    class Meta:
        model = models.Visual
        fields = ("title", "slug", "image", "settings", "caption", "published")
        widgets = {"image": forms.FileInput}


def get_visual_form(visual_type):
    try:
        return {
            constants.VisualType.BIOASSAY_AGGREGATION: EndpointAggregationForm,
            constants.VisualType.BIOASSAY_CROSSVIEW: CrossviewForm,
            constants.VisualType.ROB_HEATMAP: RoBForm,
            constants.VisualType.ROB_BARCHART: RoBForm,
            constants.VisualType.LITERATURE_TAGTREE: TagtreeForm,
            constants.VisualType.EXTERNAL_SITE: ExternalSiteForm,
            constants.VisualType.EXPLORE_HEATMAP: ExploreHeatmapForm,
            constants.VisualType.PLOTLY: PlotlyVisualForm,
            constants.VisualType.IMAGE: ImageVisualForm,
            constants.VisualType.PRISMA: PrismaVisualForm,
        }[visual_type]
    except Exception:
        raise ValueError()


class DataPivotForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        assessment = kwargs.pop("parent", None)
        super().__init__(*args, **kwargs)
        if assessment:
            self.instance.assessment = assessment
        self.helper = self.setHelper()
        self.fields["settings"].widget.attrs["rows"] = 2

    def setHelper(self):
        for fld in list(self.fields.keys()):
            widget = self.fields[fld].widget
            if type(widget) != forms.CheckboxInput:
                widget.attrs["class"] = "col-md-12"

        if self.instance.id:
            inputs = {
                "legend_text": f"Update {self.instance}",
                "help_text": "Update an existing data-pivot.",
                "cancel_url": self.instance.get_absolute_url(),
            }
        else:
            inputs = {
                "legend_text": "Create new data-pivot",
                "help_text": """
                    Create a custom-visualization for this assessment.
                    Generally, you will select a subset of available data, then
                    customize the visualization the next-page.
                """,
                "cancel_url": self.instance.get_list_url(self.instance.assessment_id),
            }
            if hasattr(self.instance, "evidence_type"):
                inputs["legend_text"] += f" ({self.instance.get_evidence_type_display()})"

        helper = BaseFormHelper(self, **inputs)

        helper.form_id = "dataPivotForm"
        return helper

    def clean_slug(self):
        return check_unique_for_assessment(self, "slug")

    def clean_title(self):
        return check_unique_for_assessment(self, "title")

    def clean_caption(self):
        caption = self.cleaned_data["caption"]
        validators.validate_hyperlinks(caption)
        return sanitize_html.clean_html(caption)


class DataPivotUploadForm(DataPivotForm):
    class Meta:
        model = models.DataPivotUpload
        exclude = ("assessment",)

    def clean(self):
        cleaned_data = super().clean()
        excel_file = cleaned_data.get("excel_file")
        worksheet_name = cleaned_data.get("worksheet_name", "")

        if excel_file:
            cannot_read = "Unable to read Excel file. Please upload an Excel file in XLSX format."

            # ensure it has correct extension
            if not excel_file.name.endswith(".xlsx"):
                self.add_error("excel_file", cannot_read)
                return

            # see if it loads
            try:
                wb = load_workbook(excel_file, read_only=True)
            except (BadZipFile, InvalidFileException):
                self.add_error("excel_file", cannot_read)
                return

            # check worksheet name
            if worksheet_name and worksheet_name not in wb.sheetnames:
                self.add_error("worksheet_name", f"Worksheet name {worksheet_name} not found.")
                return
            else:
                worksheet_name = wb.sheetnames[0]

            df = pd.read_excel(excel_file, sheet_name=worksheet_name)

            # check data
            if df.shape[0] < 2:
                self.add_error("excel_file", "Must contain at least 2 rows of data.")

            if df.shape[1] < 2:
                self.add_error("excel_file", "Must contain at least 2 columns.")


class DataPivotQueryForm(DataPivotForm):
    class Meta:
        model = models.DataPivotQuery
        fields = (
            "title",
            "slug",
            "export_style",
            "preferred_units",
            "settings",
            "caption",
            "published",
            "prefilters",
        )

    def _get_prefilter_form(self, data, **form_kwargs):
        # TODO - refactor here in in other calls; identical code
        prefix = form_kwargs.pop("prefix", None)
        prefilter = self.prefilter_cls(
            data=data, prefix=prefix, assessment=self.instance.assessment, form_kwargs=form_kwargs
        )
        form = prefilter.form
        prefilter.set_form_options(form)
        return form

    def __init__(self, *args, **kwargs):
        evidence_type = kwargs.pop("evidence_type", None)
        super().__init__(*args, **kwargs)
        if self.instance.id is None:
            self.instance.evidence_type = evidence_type

        self.prefilter_cls = prefilters.get_prefilter_cls(
            None, self.instance.evidence_type, self.instance.assessment
        )
        self.fields["prefilters"] = DynamicFormField(
            prefix="prefilters", form_class=self._get_prefilter_form, label=""
        )

        if self.instance.evidence_type == constants.StudyType.BIOASSAY:
            self.fields["preferred_units"].required = False
            self.fields["preferred_units"].choices = json.dumps(
                [
                    {"id": obj.id, "name": obj.name}
                    for obj in DoseUnits.objects.get_animal_units(self.instance.assessment)
                ]
            )
        else:
            self.fields.pop("preferred_units")

        if self.instance.evidence_type not in (
            constants.StudyType.IN_VITRO,
            constants.StudyType.BIOASSAY,
        ):
            self.fields.pop("export_style")

        self.helper = self.setHelper()

    def save(self, commit=True):
        self.instance.preferred_units = self.cleaned_data.get("preferred_units", [])
        return super().save(commit=commit)

    def clean_export_style(self):
        evidence_type = self.instance.evidence_type
        export_style = self.cleaned_data["export_style"]
        if (
            evidence_type not in (constants.StudyType.IN_VITRO, constants.StudyType.BIOASSAY)
            and export_style != constants.ExportStyle.EXPORT_GROUP
        ):
            raise forms.ValidationError(
                "Outcome/Result level export not implemented for this data-type."
            )
        return export_style


class DataPivotSettingsForm(forms.ModelForm):
    class Meta:
        model = models.DataPivot
        fields = ("settings",)


class DataPivotSelectorForm(CopyForm):
    legend_text = "Copy data pivot"
    help_text = """
        Select an existing data pivot and copy as a new data pivot. This includes all
        model-settings, and the selected dataset. You will be taken to a new view to
        create a new data pivot, but the form will be pre-populated using the values from
        the currently-selected data pivot."""
    create_url_pattern = "summary:visualization_create"
    selector = forms.ModelChoiceField(
        queryset=models.DataPivot.objects.all(), empty_label=None, label="Select template"
    )
    reset_row_overrides = forms.BooleanField(
        help_text="Reset all row-level customization in the data-pivot copy",
        required=False,
        initial=True,
    )

    def __init__(self, *args, **kw):
        user = kw.pop("user")
        super().__init__(*args, **kw)
        self.fields["selector"].queryset = (
            models.DataPivot.objects.clonable_queryset(user)
            .select_related("assessment")
            .order_by("assessment", "title")
        )
        self.fields["selector"].label_from_instance = lambda obj: f"{obj.assessment} | {obj}"

    def get_success_url(self):
        dp = self.cleaned_data["selector"]
        reset_row_overrides = self.cleaned_data["reset_row_overrides"]

        if hasattr(dp, "datapivotupload"):
            url = reverse("summary:dp_new-file", args=(self.parent.id,))
        else:
            url = reverse(
                "summary:dp_new-query", args=(self.parent.id, dp.datapivotquery.evidence_type)
            )

        url += f"?initial={dp.pk}"

        if reset_row_overrides:
            url += "&reset_row_overrides=1"

        return url

    def get_cancel_url(self):
        return reverse("summary:visualization_list", args=(self.parent.id,))


class SmartTagForm(forms.Form):
    RESOURCE_CHOICES = (
        ("study", "Study"),
        ("endpoint", "Endpoint"),
        ("visual", "Visualization"),
        ("data_pivot", "Data Pivot"),
    )
    resource = forms.ChoiceField(choices=RESOURCE_CHOICES)
    study = AutocompleteChoiceField(
        autocomplete_class=StudyAutocomplete,
        help_text="Type a few characters of the study name, then click to select.",
    )
    endpoint = AutocompleteChoiceField(
        autocomplete_class=EndpointAutocomplete,
        help_text="Type a few characters of the endpoint name, then click to select.",
    )
    visual = AutocompleteChoiceField(
        autocomplete_class=autocomplete.VisualAutocomplete,
        help_text="Type a few characters of the visual name, then click to select.",
    )
    data_pivot = AutocompleteChoiceField(
        autocomplete_class=autocomplete.DataPivotAutocomplete,
        help_text="Type a few characters of the data-pivot name, then click to select.",
    )

    def __init__(self, *args, **kwargs):
        assessment_id = kwargs.pop("assessment_id", -1)
        super().__init__(*args, **kwargs)
        self.fields["study"].set_filters({"assessment_id": assessment_id})
        self.fields["endpoint"].set_filters(
            {"animal_group__experiment__study__assessment_id": assessment_id}
        )
        self.fields["visual"].set_filters({"assessment_id": assessment_id})
        self.fields["data_pivot"].set_filters({"assessment_id": assessment_id})
