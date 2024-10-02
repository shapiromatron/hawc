import pytest
from openpyxl import load_workbook

from hawc.apps.assessment.models import Assessment
from hawc.apps.epiv2.exports import EpiFlatComplete, tabular_export
from hawc.apps.epiv2.models import DataExtraction


@pytest.mark.django_db
class TestEpiFlatComplete:
    def test_export(self):
        assessment = Assessment.objects.get(id=1)
        qs = DataExtraction.objects.filter(design__study__assessment_id=1)
        assert qs.count() > 0
        exporter = EpiFlatComplete(qs, assessment=assessment, filename="fn")
        df = exporter.build_df()
        assert df.shape == (12, 118)


@pytest.mark.django_db
def test_tabular_export():
    f = tabular_export(1, False)
    wb = load_workbook(f)
    assert len(wb.sheetnames) == 7
    assert wb["design"].max_row == 3
    assert wb["design"].max_column == 42
